"""Full end-to-end test of CSV converter with 1644 trades"""
import sys
import os
from pathlib import Path
import time
import threading

# Add v736 to path for imports
sys.path.insert(0, r'c:\Users\LENOVO THINKPAD\Documents\Aventa_AI_2026\Aventa_HFT_Pro_2026_v736')

from csv_to_excel_converter_gui_736 import CSVToExcelConverter
import tkinter as tk
import pandas as pd

print('[E2E TEST] Full converter end-to-end test with 1644 trades')
print('=' * 60)

# Create root window
root = tk.Tk()
root.withdraw()

# Initialize converter
try:
    converter = CSVToExcelConverter(root)
    print('[OK] Converter initialized')
except Exception as e:
    print(f'[ERROR] Failed to initialize: {e}')
    import traceback
    traceback.print_exc()
    sys.exit(1)

# Set input file and output folder
csv_file = 'test_1644_trades.csv'
output_folder = 'test_output'

if not os.path.exists(csv_file):
    print(f'[ERROR] CSV file not found: {csv_file}')
    sys.exit(1)

# Check data
df = pd.read_csv(csv_file)
print(f'[OK] CSV loaded: {len(df)} trades, {len(df.columns)} columns')

# Set converter variables
converter.csv_file_var.set(csv_file)
converter.output_folder_var.set(output_folder)

# Create output folder
os.makedirs(output_folder, exist_ok=True)

print('\n[PROCESS] Starting conversion...')
print('-' * 60)

# Clear previous logs
converter.log_text.delete(1.0, tk.END)

# Track if conversion is done
conversion_done = False

def monitor_conversion():
    global conversion_done
    # Keep GUI responsive and wait for conversion
    for i in range(300):  # Wait up to 30 seconds
        root.update()
        time.sleep(0.1)
        
        # Check if output files have been created
        output_files = list(Path(output_folder).glob('*.xlsx'))
        if output_files and not converter.is_converting:
            conversion_done = True
            break
    
    root.quit()

# Start monitoring thread
monitor_thread = threading.Thread(target=monitor_conversion)
monitor_thread.daemon = True
monitor_thread.start()

try:
    # Start conversion
    converter.convert_file()
    
    # Run event loop
    root.mainloop()
    
    # Wait a moment for files to be written
    time.sleep(1)
    
    # Check if conversion succeeded
    output_files = list(Path(output_folder).glob('*.xlsx'))
    
    if output_files:
        xlsx_file = output_files[0]
        file_size = xlsx_file.stat().st_size / 1024
        print(f'[OK] Excel file created: {xlsx_file.name} ({file_size:.1f} KB)')
        
        # Verify all sheets
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_file)
        visible_sheets = [name for name in wb.sheetnames 
                         if wb[name].sheet_state == 'visible']
        
        print(f'[OK] Sheets created: {len(wb.sheetnames)}')
        print(f'[OK] Sheets visible: {len(visible_sheets)}')
        print(f'[OK] Sheet list: {wb.sheetnames}')
        
        # Check for PDF
        pdf_files = list(Path(output_folder).glob('*.pdf'))
        if pdf_files:
            pdf_file = pdf_files[0]
            pdf_size = pdf_file.stat().st_size / 1024
            print(f'[OK] PDF file created: {pdf_file.name} ({pdf_size:.1f} KB)')
        
        # Get logs
        log_output = converter.log_text.get(1.0, tk.END)
        if log_output.strip():
            print('\n[LOGS]')
            for line in log_output.split('\n')[-20:]:
                if line.strip():
                    print(f'  {line}')
        
        print('\n' + '=' * 60)
        print('[SUCCESS] E2E TEST PASSED: Full conversion successful!')
        print('=' * 60)
        
    else:
        print('[ERROR] No Excel files generated')
        log_output = converter.log_text.get(1.0, tk.END)
        if log_output.strip():
            print('\n[LOGS]')
            print(log_output)
        sys.exit(1)
        
except Exception as e:
    print(f'[ERROR] Test failed: {e}')
    import traceback
    traceback.print_exc()
    sys.exit(1)

finally:
    try:
        root.destroy()
    except:
        pass
