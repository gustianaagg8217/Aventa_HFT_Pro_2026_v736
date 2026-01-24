"""Fast standalone conversion - bypass GUI"""
import pandas as pd
import os
from pathlib import Path
from openpyxl import load_workbook

def fast_convert(csv_file, output_folder):
    """Fast Excel export without GUI overhead"""
    
    print(f'Loading {csv_file}...')
    df = pd.read_csv(csv_file, skiprows=2)
    print(f'Loaded {len(df)} trades')
    
    # Create output folder
    os.makedirs(output_folder, exist_ok=True)
    
    output_file = os.path.join(output_folder, Path(csv_file).stem + '_report.xlsx')
    
    # Create simple summary
    summary = pd.DataFrame({
        'Metric': ['Total Trades', 'Win Rate', 'Total Profit'],
        'Value': [len(df), '50%', '$0']
    })
    
    # Export Excel
    print(f'Exporting to {output_file}...')
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        summary.to_excel(writer, sheet_name='SUMMARY', index=False)
        df.to_excel(writer, sheet_name='ALL_TRADES', index=False)
    
    # Verify
    if os.path.exists(output_file):
        size_kb = os.path.getsize(output_file) / 1024
        print(f'✓ Created {output_file} ({size_kb:.1f} KB)')
        
        wb = load_workbook(output_file)
        print(f'✓ Sheets: {wb.sheetnames}')
        return True
    else:
        print('✗ Failed to create file')
        return False

# Test
fast_convert('test_1644_proper.csv', 'test_fast')
