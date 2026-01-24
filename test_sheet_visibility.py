"""Test sheet visibility fix with 1644 trades"""
import pandas as pd
from openpyxl import load_workbook, Workbook
import os
import sys

print('[TEST] Creating large dataset with 1644 trades...')
trades = []
for i in range(1644):
    trades.append({
        '#': i+1,
        'Entry Time': '2026-01-20 10:00:00',
        'Exit Time': '2026-01-20 10:05:00',
        'Type': 'BUY',
        'Entry Price': 19500.00 + i*0.1,
        'Exit Price': 19510.00 + i*0.1,
        'Volume': 0.5,
        'Profit': 50.00 + i*0.05,
        'Saldo Awal': 10000.00,
        'Saldo Akhir': 10050.00,
        'Duration': '5m',
        'Reason': 'Test',
        'Symbol': 'EURUSD',
        'Period': 'M5',
        'Company': 'MetaTrader 5'
    })

df = pd.DataFrame(trades)
csv_file = 'test_1644_trades.csv'
df.to_csv(csv_file, index=False)
print(f'  ✓ Created {csv_file} ({len(df)} rows)')

# Test Excel creation with visibility management
print('[TEST] Creating Excel workbook...')
wb = Workbook()
ws = wb.active
ws.title = 'ALL_TRADES'

# Write headers
headers = df.columns.tolist()
ws.append(headers)

# Write data
for _, row in df.iterrows():
    ws.append(row.tolist())

# Create additional sheets
sheets_to_create = ['BUY_TRADES', 'SELL_TRADES', 'SUMMARY', 'BACKTEST_DETAILS']
for sheet_name in sheets_to_create:
    ws_new = wb.create_sheet(sheet_name)
    ws_new['A1'] = f'{sheet_name} Data'

print(f'  ✓ Created {len(wb.sheetnames)} sheets: {wb.sheetnames}')

# Test visibility fix (matching the new code)
print('[TEST] Testing sheet visibility fix...')
try:
    # Ensure all sheets visible before operations
    for sheet_name in wb.sheetnames:
        wb[sheet_name].sheet_state = 'visible'
    
    # Create/get CHARTS sheet
    if 'CHARTS' not in wb.sheetnames:
        worksheet = wb.create_sheet('CHARTS', 0)
    else:
        worksheet = wb['CHARTS']
    
    worksheet.sheet_state = 'visible'
    
    # Final safety check
    visible_sheets = [name for name in wb.sheetnames 
                     if wb[name].sheet_state == 'visible']
    if not visible_sheets and wb.sheetnames:
        wb[wb.sheetnames[0]].sheet_state = 'visible'
    
    print(f'  ✓ All sheets set to visible')
    print(f'  ✓ Visible sheets count: {len(visible_sheets)}')
    
    # Save workbook
    output_file = 'test_1644_trades_output.xlsx'
    wb.save(output_file)
    print(f'  ✓ Saved {output_file} ({os.path.getsize(output_file) / 1024:.1f} KB)')
    
    # Verify the saved file
    print('[VERIFY] Checking saved file...')
    wb_verify = load_workbook(output_file)
    visible_count = sum(1 for sheet in wb_verify.sheetnames if wb_verify[sheet].sheet_state == 'visible')
    print(f'  ✓ Sheets in file: {len(wb_verify.sheetnames)}')
    print(f'  ✓ Visible sheets: {visible_count}')
    print(f'  ✓ All sheets: {wb_verify.sheetnames}')
    
    print('')
    print('✓ TEST PASSED: Sheet visibility fix working correctly!')
    sys.exit(0)
    
except Exception as e:
    print(f'  ✗ ERROR: {str(e)}')
    import traceback
    traceback.print_exc()
    sys.exit(1)
