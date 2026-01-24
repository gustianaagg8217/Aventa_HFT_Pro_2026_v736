"""Full end-to-end test of CSV converter with 1644 trades"""
import sys
import os
from pathlib import Path

# Add v736 to path for imports
sys.path.insert(0, r'c:\Users\LENOVO THINKPAD\Documents\Aventa_AI_2026\Aventa_HFT_Pro_2026_v736')

from csv_to_excel_converter_gui_736 import CSVToExcelConverter
import tkinter as tk
import pandas as pd

print('[E2E TEST] Full converter end-to-end test with 1644 trades')
print('=' * 60)

# Create root window (hidden)
root = tk.Tk()
root.withdraw()

# Initialize converter
try:
    converter = CSVToExcelConverter(root)
    print('[OK] Converter initialized')
except Exception as e:
    print(f'[ERROR] Failed to initialize: {e}')
    sys.exit(1)

# Set input file and output folder
csv_file = 'test_1644_trades.csv'
output_folder = 'test_output'

if not os.path.exists(csv_file):
    print(f'[ERROR] CSV file not found: {csv_file}')
    sys.exit(1)

# Check data
df = pd.read_csv(csv_file)
print(f'[OK] CSV loaded: {len(df)} trades, {len(df.columns)} columns')

# Set converter variables
converter.csv_file_var.set(csv_file)
converter.output_folder_var.set(output_folder)

# Create output folder
os.makedirs(output_folder, exist_ok=True)

print('\n[PROCESS] Starting conversion...')
print('-' * 60)

# Clear previous logs
converter.log_text.delete(1.0, tk.END)

try:
    # Call the conversion method
    converter.convert_file()
    
    # Wait a moment for async operations
    root.update()
    
    # Check if conversion succeeded (by checking output folder)
    output_files = list(Path(output_folder).glob('*.xlsx'))
    
    if output_files:
        xlsx_file = output_files[0]
        file_size = xlsx_file.stat().st_size / 1024
        print(f'[OK] Excel file created: {xlsx_file.name} ({file_size:.1f} KB)')
        
        # Verify all sheets
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_file)
        visible_sheets = [name for name in wb.sheetnames 
                         if wb[name].sheet_state == 'visible']
        
        print(f'[OK] Sheets created: {len(wb.sheetnames)}')
        print(f'[OK] Sheets visible: {len(visible_sheets)}')
        print(f'[OK] Sheet list: {wb.sheetnames}')
        
        # Check for PDF
        pdf_files = list(Path(output_folder).glob('*.pdf'))
        if pdf_files:
            pdf_file = pdf_files[0]
            pdf_size = pdf_file.stat().st_size / 1024
            print(f'[OK] PDF file created: {pdf_file.name} ({pdf_size:.1f} KB)')
        
        print('\n' + '=' * 60)
        print('[SUCCESS] E2E TEST PASSED: Full conversion successful!')
        print('=' * 60)
        
    else:
        print('[ERROR] No Excel files generated')
        print('Log output:')
        print(converter.log_text.get(1.0, tk.END))
        sys.exit(1)
        
except Exception as e:
    print(f'[ERROR] Conversion failed: {e}')
    import traceback
    traceback.print_exc()
    print('\nLog output:')
    print(converter.log_text.get(1.0, tk.END))
    sys.exit(1)

finally:
    root.destroy()
