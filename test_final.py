"""Direct test - capture logs properly"""
import sys
import os
sys.path.insert(0, r'c:\Users\LENOVO THINKPAD\Documents\Aventa_AI_2026\Aventa_HFT_Pro_2026_v736')

from csv_to_excel_converter_gui_736 import CSVToExcelConverter
import tkinter as tk
from pathlib import Path
import io
from contextlib import redirect_stdout, redirect_stderr

print('[FINAL TEST] Direct converter test with proper logging')
print('=' * 60)

# Setup
root = tk.Tk()
root.withdraw()

# Create converter
converter = CSVToExcelConverter(root)

# Setup paths
csv_file = 'test_1644_proper.csv'
output_folder = 'test_final_direct'
os.makedirs(output_folder, exist_ok=True)

converter.csv_file_var.set(csv_file)
converter.output_folder_var.set(output_folder)

# Run conversion
print('Running conversion...')
try:
    converter._do_conversion()
    print('Conversion finished')
except Exception as e:
    print(f'Exception during conversion: {e}')
    import traceback
    traceback.print_exc()

# Process any pending events
root.update()

# Get and print logs
log_text = converter.log_text.get(1.0, tk.END)
print('[CONVERTER LOGS]')
for line in log_text.split('\n')[-30:]:
    if line.strip():
        try:
            print(f'  {line}')
        except:
            print(f'  [unprintable line]')

# Check output
print('\n[OUTPUT FILES]')
xlsx_files = list(Path(output_folder).glob('*.xlsx'))
pdf_files = list(Path(output_folder).glob('*.pdf'))

for f in xlsx_files:
    size_kb = f.stat().st_size / 1024
    print(f'  XLSX: {f.name} ({size_kb:.1f} KB)')

for f in pdf_files:
    size_kb = f.stat().st_size / 1024
    print(f'  PDF: {f.name} ({size_kb:.1f} KB)')

if not (xlsx_files or pdf_files):
    print('  (No files created)')

# Verify Excel file if created
if xlsx_files:
    try:
        from openpyxl import load_workbook
        xlsx_file = xlsx_files[0]
        wb = load_workbook(xlsx_file)
        visible_count = sum(1 for s in wb.sheetnames if wb[s].sheet_state == 'visible')
        print(f'\n[VERIFICATION]')
        print(f'  Sheets: {len(wb.sheetnames)} ({visible_count} visible)')
        print(f'  Sheet names: {wb.sheetnames}')
        if wb.sheetnames:
            for sheet_name in wb.sheetnames[:3]:  # Show first 3 sheets
                ws = wb[sheet_name]
                print(f'    {sheet_name}: {ws.dimensions}')
    except Exception as e:
        print(f'[ERROR] Could not verify: {e}')

root.destroy()

print('\n' + '=' * 60)
if xlsx_files and xlsx_files[0].stat().st_size > 10000:  # Should be > 10KB
    print('[SUCCESS] Conversion test PASSED!')
else:
    print('[ERROR] File too small or missing!')
    sys.exit(1)
