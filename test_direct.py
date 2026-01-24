"""Direct test of the _do_conversion logic"""
import sys
import os
sys.path.insert(0, r'c:\Users\LENOVO THINKPAD\Documents\Aventa_AI_2026\Aventa_HFT_Pro_2026_v736')

from csv_to_excel_converter_gui_736 import CSVToExcelConverter
import tkinter as tk
from pathlib import Path

print('[TEST] Direct conversion test with 1644 trades')
print('=' * 60)

# Setup
root = tk.Tk()
root.withdraw()

converter = CSVToExcelConverter(root)

# Configure paths
csv_file = 'test_1644_proper.csv'
output_folder = 'test_output_final'

if not os.path.exists(csv_file):
    print(f'[ERROR] CSV not found: {csv_file}')
    sys.exit(1)

# Create output folder
os.makedirs(output_folder, exist_ok=True)

# Set variables
converter.csv_file_var.set(csv_file)
converter.output_folder_var.set(output_folder)

# Clear any previous logs
converter.log_text.delete(1.0, tk.END)

print('[RUN] Calling _do_conversion directly...')
print('-' * 60)

try:
    # Call conversion directly (no threading)
    converter._do_conversion()
    
    # Get logs
    log_output = converter.log_text.get(1.0, tk.END)
    print(log_output)
    
    # Check output
    print('-' * 60)
    xlsx_files = list(Path(output_folder).glob('*.xlsx'))
    
    if xlsx_files:
        for f in xlsx_files:
            size_kb = f.stat().st_size / 1024
            print(f'[OK] Excel: {f.name} ({size_kb:.1f} KB)')
            
            # Verify sheets
            from openpyxl import load_workbook
            wb = load_workbook(f)
            visible = sum(1 for s in wb.sheetnames if wb[s].sheet_state == 'visible')
            print(f'[OK] Sheets: {len(wb.sheetnames)} total, {visible} visible')
            print(f'[OK] Sheet names: {wb.sheetnames}')
    
    pdf_files = list(Path(output_folder).glob('*.pdf'))
    if pdf_files:
        for f in pdf_files:
            size_kb = f.stat().st_size / 1024
            print(f'[OK] PDF: {f.name} ({size_kb:.1f} KB)')
    
    if xlsx_files or pdf_files:
        print('\n[SUCCESS] Conversion completed!')
    else:
        print('\n[ERROR] No output files created')
        sys.exit(1)
        
except Exception as e:
    print(f'[ERROR] {e}')
    import traceback
    traceback.print_exc()
    sys.exit(1)

finally:
    root.destroy()
