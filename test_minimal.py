"""Minimal test - just call conversion core logic"""
import sys
import os
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path

sys.path.insert(0, r'c:\Users\LENOVO THINKPAD\Documents\Aventa_AI_2026\Aventa_HFT_Pro_2026_v736')

print('[MINIMAL TEST] Testing ExcelWriter visibility fix')
print('=' * 60)

# Simple Excel export using pandas ExcelWriter
csv_file = 'test_1644_proper.csv'
output_folder = 'test_minimal'
os.makedirs(output_folder, exist_ok=True)

# Load CSV
df = pd.read_csv(csv_file, skiprows=2)
print(f'Loaded: {len(df)} trades')

# Create data for export
output_file = os.path.join(output_folder, 'minimal_test.xlsx')

print('Creating workbook...')
try:
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        # Set visibility BEFORE exporting
        if hasattr(writer.book, 'worksheets'):
            for ws in writer.book.worksheets:
                ws.sheet_state = 'visible'
        
        # Export sheets
        df.to_excel(writer, sheet_name="ALL_TRADES", index=False)
        print('  Exported ALL_TRADES')
        
        # Create summary
        summary_df = pd.DataFrame({
            'Metric': ['Total Trades', 'Win Rate', 'Total Profit'],
            'Value': [len(df), '50%', '$0.00']
        })
        summary_df.to_excel(writer, sheet_name="SUMMARY", index=False)
        print('  Exported SUMMARY')
        
        # Ensure visibility AFTER all exports
        for ws in writer.book.worksheets:
            ws.sheet_state = 'visible'
            print(f'  Set {ws.title} to visible')
    
    print('Workbook saved')
    
    # Verify
    print('Verifying...')
    wb = load_workbook(output_file)
    print(f'  Sheets: {wb.sheetnames}')
    visible_count = sum(1 for s in wb.sheetnames if wb[s].sheet_state == 'visible')
    print(f'  Visible: {visible_count}/{len(wb.sheetnames)}')
    
    if len(wb.sheetnames) > 0 and visible_count > 0:
        print('\n[SUCCESS] Minimal test passed!')
    else:
        print('\n[ERROR] No sheets or all hidden!')
        sys.exit(1)
        
except Exception as e:
    print(f'[ERROR] {e}')
    import traceback
    traceback.print_exc()
    sys.exit(1)
