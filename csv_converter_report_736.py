import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as XLImage
import threading
from datetime import datetime
try:
    from fpdf import FPDF
except ImportError:
    FPDF = None
try:
    import matplotlib.pyplot as plt
    import matplotlib
    matplotlib.use('Agg')  # Use non-GUI backend
except ImportError:
    matplotlib = None

class CSVToExcelConverter:
    def __init__(self, root):
        self.root = root
        self.root.title("CSV to Excel Converter Pro")
        self.root.geometry("700x600")
        self.root.resizable(False, False)
        
        # Apply dark theme
        self.root.configure(bg='#1a1e3a')
        
        # Variables
        self.csv_file_var = tk.StringVar()
        self.output_folder_var = tk.StringVar(value="Backtest")
        self.is_converting = False
        
        # Configure styles
        self.setup_styles()
        self.setup_ui()
    
    
    def setup_styles(self):
        """Setup dark theme styles to match main GUI"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Dark theme colors
        bg_color = '#1a1e3a'
        fg_color = '#e0e0e0'
        accent_color = '#00e676'
        warning_color = '#ffd600'
        error_color = '#ff1744'
        
        # Configure colors for ttk widgets
        style.configure('TFrame', background=bg_color)
        style.configure('TLabel', background=bg_color, foreground=fg_color)
        style.configure('TLabelframe', background=bg_color, foreground=fg_color)
        style.configure('TLabelframe.Label', background=bg_color, foreground=fg_color)
        style.configure('TButton', background=bg_color, foreground=fg_color, borderwidth=1)
        style.map('TButton', 
                  background=[('active', '#2a3f5f')],
                  foreground=[('active', accent_color)])
        style.configure('TEntry', fieldbackground='#0f1219', foreground=fg_color, borderwidth=1)
        style.configure('TCheckbutton', background=bg_color, foreground=fg_color)
        
    def setup_ui(self):
        """Setup UI components"""
        # Main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # === CSV FILE SELECTION ===
        file_frame = ttk.LabelFrame(main_frame, text="📁 CSV File Selection", padding="10")
        file_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(file_frame, text="CSV File:", font=("Segoe UI", 10)).pack(anchor=tk.W, pady=5)
        
        file_input_frame = ttk.Frame(file_frame)
        file_input_frame.pack(fill=tk.X, pady=5)
        
        ttk.Entry(file_input_frame, textvariable=self.csv_file_var, width=50).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        ttk.Button(file_input_frame, text="🔍 Browse", command=self.browse_csv, width=12).pack(side=tk.LEFT)
        
        # === OUTPUT FOLDER ===
        output_frame = ttk.LabelFrame(main_frame, text="📂 Output Folder", padding="10")
        output_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(output_frame, text="Output Folder:", font=("Segoe UI", 10)).pack(anchor=tk.W, pady=5)
        
        output_input_frame = ttk.Frame(output_frame)
        output_input_frame.pack(fill=tk.X, pady=5)
        
        ttk.Entry(output_input_frame, textvariable=self.output_folder_var, width=50).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        ttk.Button(output_input_frame, text="🔍 Browse", command=self.browse_output, width=12).pack(side=tk.LEFT)
        
        # === CONTROL BUTTONS ===
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.convert_btn = ttk.Button(button_frame, text="🚀 Convert to Excel", command=self.convert_file, width=25)
        self.convert_btn.pack(side=tk.LEFT, padx=5)
        
        self.pdf_btn = ttk.Button(button_frame, text="📄 Convert to PDF", command=self.convert_to_pdf, width=20)
        self.pdf_btn.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(button_frame, text="🗑️ Clear", command=self.clear_fields, width=12).pack(side=tk.LEFT, padx=5)
        
        # === LOG AREA ===
        log_frame = ttk.LabelFrame(main_frame, text="📝 Conversion Log", padding="5")
        log_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        self.log_text = scrolledtext.ScrolledText(log_frame, wrap=tk.WORD, height=15, 
                                                  bg='#0f1219', fg='#e0e0e0', font=("Courier", 9),
                                                  insertbackground='#00e676')
        self.log_text.pack(fill=tk.BOTH, expand=True)
        
        self.log_text.tag_config("INFO", foreground="#00e676")
        self.log_text.tag_config("SUCCESS", foreground="#00b0ff")
        self.log_text.tag_config("ERROR", foreground="#ff1744")
        self.log_text.tag_config("WARNING", foreground="#ffd600")
        
        # === STATUS BAR ===
        self.status_var = tk.StringVar(value="Ready")
        status_bar = ttk.Label(main_frame, textvariable=self.status_var, relief=tk.SUNKEN, font=("Segoe UI", 9))
        status_bar.pack(fill=tk.X)
    
    def _add_market_column(self, df):
        """
        Add Market column based on trading time (WIB - Waktu Indonesia Barat)
        Asia Session: 05:00 - 15:00 WIB
        London Session: 15:00 - 23:30 WIB
        New York Session: 20:00 - 04:00 WIB (includes early morning hours 00:00-04:59)
        """
        try:
            # Find time column (could be 'Entry Time', 'Time', 'entry time', etc.)
            time_col = None
            for col in df.columns:
                col_lower = col.lower()
                if 'entry' in col_lower and 'time' in col_lower:
                    time_col = col
                    break
            
            if time_col is None:
                for col in df.columns:
                    col_lower = col.lower()
                    if 'time' in col_lower:
                        time_col = col
                        break
            
            if time_col is None:
                # If no time column found, mark all as Unknown
                df['Market'] = 'Unknown'
                return df
            
            def get_market(time_str):
                try:
                    # Parse time (handle various formats)
                    if isinstance(time_str, str):
                        # Format: "2025-12-22 01:50:00" or "01:50:00" or "01:50"
                        time_parts = time_str.strip().split()
                        
                        # Get the time part (last element or only element)
                        time_part = time_parts[-1]
                        time_values = time_part.split(':')
                        
                        if len(time_values) >= 1:
                            hour = int(time_values[0])
                        else:
                            return 'Unknown'
                    else:
                        # If it's a datetime object
                        hour = int(time_str.hour) if hasattr(time_str, 'hour') else 0
                    
                    # Determine market based on hour (WIB time)
                    # New York: 20:00 - 04:00 (includes early morning)
                    # Asia: 05:00 - 15:00 (before London)
                    # London: 15:00 - 23:30 (can overlap with NY)
                    
                    if 5 <= hour < 15:
                        return 'Asia'
                    elif 15 <= hour < 20:
                        return 'London'
                    elif hour >= 20 or hour < 5:
                        # 20:00-23:59 and 00:00-04:59
                        return 'New York'
                    else:
                        return 'Unknown'
                except Exception as e:
                    return 'Unknown'
            
            df['Market'] = df[time_col].apply(get_market)
            return df
        
        except Exception as e:
            self.add_log(f"⚠️ Error adding market column: {e}", "WARNING")
            df['Market'] = 'Unknown'
            return df
    
    def browse_csv(self):
        """Browse for CSV file"""
        filename = filedialog.askopenfilename(
            title="Select CSV File",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            initialdir="Backtest"
        )
        if filename:
            self.csv_file_var.set(filename)
            self.add_log(f"Selected file: {filename}", "INFO")
    
    def browse_output(self):
        """Browse for output folder"""
        folder = filedialog.askdirectory(title="Select Output Folder", initialdir="Backtest")
        if folder:
            self.output_folder_var.set(folder)
            self.add_log(f"Output folder: {folder}", "INFO")
    
    def clear_fields(self):
        """Clear all fields"""
        self.csv_file_var.set("")
        self.output_folder_var.set("Backtest")
        self.log_text.delete(1.0, tk.END)
        self.status_var.set("Ready")
        self.add_log("Fields cleared", "INFO")
    
    def add_log(self, message, level="INFO"):
        """Add message to log"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_entry = f"[{timestamp}] {message}\n"
        self.log_text.insert(tk.END, log_entry, level)
        self.log_text.see(tk.END)
        self.root.update()
    
    def convert_file(self):
        """Convert CSV to Excel in background thread"""
        if not self.csv_file_var.get():
            messagebox.showerror("Error", "Please select a CSV file first!")
            return
        
        if not os.path.exists(self.csv_file_var.get()):
            messagebox.showerror("Error", f"File not found: {self.csv_file_var.get()}")
            return
        
        # Create output folder if not exists
        output_folder = self.output_folder_var.get()
        if not os.path.exists(output_folder):
            try:
                os.makedirs(output_folder)
                self.add_log(f"Created output folder: {output_folder}", "INFO")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to create output folder: {e}")
                return
        
        # Disable convert button and start conversion in thread
        self.convert_btn.config(state=tk.DISABLED)
        self.is_converting = True
        self.status_var.set("Converting...")
        
        thread = threading.Thread(target=self._do_conversion, daemon=True)
        thread.start()
    
    def _do_conversion(self):
        """Actual conversion logic (runs in background thread)"""
        try:
            csv_file = self.csv_file_var.get()
            output_folder = self.output_folder_var.get()
            
            self.add_log("=" * 60, "INFO")
            self.add_log("🚀 Starting conversion...", "INFO")
            self.add_log(f"Input: {csv_file}", "INFO")
            self.add_log(f"Output folder: {output_folder}", "INFO")
            
            # Load CSV
            self.add_log("📖 Loading CSV file...", "INFO")
            df = pd.read_csv(csv_file, skiprows=2)
            self.add_log(f"✓ Loaded {len(df)} trades", "SUCCESS")
            
            # Add Market column based on time
            self.add_log("🌍 Detecting trading markets...", "INFO")
            df = self._add_market_column(df)
            self.add_log("✓ Market column added", "SUCCESS")
            
            # Get initial balance
            self.add_log("💰 Extracting initial balance...", "INFO")
            initial_balance = 0.0
            try:
                with open(csv_file, 'r') as f:
                    first_line = f.readline().strip()
                    if 'Initial Balance' in first_line:
                        parts = first_line.split(',')
                        if len(parts) > 1:
                            initial_balance = float(parts[1].strip())
            except Exception as e:
                self.add_log(f"⚠️ Could not extract initial balance: {e}", "WARNING")
            
            self.add_log(f"✓ Initial Balance: ${initial_balance:.2f}", "SUCCESS")
            
            # Calculate metrics
            self.add_log("📊 Calculating metrics...", "INFO")
            
            COL_PROFIT = "Profit"
            COL_TYPE = "Type"
            COL_MARKET = "Market"
            COL_TIME = "Time"  # Assuming CSV has Time column
            
            buy_df = df[df[COL_TYPE].str.upper() == "BUY"]
            sell_df = df[df[COL_TYPE].str.upper() == "SELL"]
            
            total_trades = len(df)
            total_profit = df[COL_PROFIT].sum()
            
            # Calculate final balance (closing balance)
            closing_balance = initial_balance + total_profit
            
            # Get last balance for Slado Akhir
            slado_akhir = closing_balance
            
            # Market statistics
            market_stats = {}
            if COL_MARKET in df.columns:
                for market in df[COL_MARKET].unique():
                    market_df = df[df[COL_MARKET] == market]
                    market_trades = len(market_df)
                    market_profit = market_df[COL_PROFIT].sum()
                    market_wins = len(market_df[market_df[COL_PROFIT] > 0])
                    market_winrate = (market_wins / market_trades * 100) if market_trades > 0 else 0
                    
                    market_stats[market] = {
                        'trades': market_trades,
                        'profit': market_profit,
                        'wins': market_wins,
                        'winrate': market_winrate
                    }
            
            win_trades = df[df[COL_PROFIT] > 0]
            loss_trades = df[df[COL_PROFIT] < 0]
            
            win_count = len(win_trades)
            loss_count = len(loss_trades)
            winrate = (win_count / total_trades) * 100 if total_trades else 0
            
            avg_profit = df[COL_PROFIT].mean()
            max_profit = df[COL_PROFIT].max()
            max_loss = df[COL_PROFIT].min()
            
            gross_profit = win_trades[COL_PROFIT].sum()
            gross_loss = abs(loss_trades[COL_PROFIT].sum())
            profit_factor = gross_profit / gross_loss if gross_loss != 0 else 0
            
            # Calculate advanced metrics
            net_profit = total_profit
            return_percent = ((closing_balance - initial_balance) / initial_balance * 100) if initial_balance > 0 else 0
            
            # Drawdown calculations
            cumulative_profit = df[COL_PROFIT].cumsum() + initial_balance
            running_max = cumulative_profit.expanding().max()
            drawdown_absolute = running_max - cumulative_profit
            max_drawdown_absolute = drawdown_absolute.max()
            max_drawdown_relative = (max_drawdown_absolute / running_max.max() * 100) if running_max.max() > 0 else 0
            
            # Recovery factor (Net Profit / Max Drawdown)
            recovery_factor = net_profit / max_drawdown_absolute if max_drawdown_absolute > 0 else 0
            
            # AHPR (Average Holding Period Return)
            ahpr = ((closing_balance / initial_balance) ** (1 / total_trades) - 1) * 100 if total_trades > 0 and initial_balance > 0 else 0
            
            self.add_log(f"✓ Total Trades: {total_trades}", "INFO")
            self.add_log(f"✓ Win Rate: {winrate:.1f}%", "INFO")
            self.add_log(f"✓ Total Profit: ${total_profit:.2f}", "INFO")
            
            # Generate output filename
            self.add_log("📝 Generating output filename...", "INFO")
            input_basename = os.path.basename(os.path.splitext(csv_file)[0])
            output_file = os.path.join(output_folder, input_basename + "_report.xlsx")
            
            # Export to Excel
            self.add_log("📤 Exporting to Excel...", "INFO")
            
            # Generate performance charts for Excel
            self.add_log("📊 Generating performance charts...", "INFO")
            chart_images = self._generate_performance_charts(df, output_folder)
            
            # Build summary data first
            summary_df = pd.DataFrame({
                "Metric": [
                    "Initial Balance ($)",
                    "Total Trades",
                    "Total Profit",
                    "Closing Balance / Slado Akhir ($)",
                    "Win Trades",
                    "Loss Trades",
                    "Winrate (%)",
                    "Average Profit",
                    "Max Profit",
                    "Max Loss",
                    "Profit Factor",
                    "",
                    "MARKET PERFORMANCE",
                    ""
                ],
                "Value": [
                    round(initial_balance, 2),
                    total_trades,
                    round(total_profit, 2),
                    round(slado_akhir, 2),
                    win_count,
                    loss_count,
                    round(winrate, 2),
                    round(avg_profit, 2),
                    round(max_profit, 2),
                    round(max_loss, 2),
                    round(profit_factor, 2),
                    "",
                    "",
                    ""
                ]
            })
            
            # Add market statistics to summary
            if market_stats:
                market_rows = []
                for market, stats in sorted(market_stats.items()):
                    market_rows.append({
                        "Metric": f"  {market} - Trades",
                        "Value": stats['trades']
                    })
                    market_rows.append({
                        "Metric": f"  {market} - Profit",
                        "Value": round(stats['profit'], 2)
                    })
                    market_rows.append({
                        "Metric": f"  {market} - Winrate (%)",
                        "Value": round(stats['winrate'], 2)
                    })
                    market_rows.append({
                        "Metric": "",
                        "Value": ""
                    })
                
                market_df = pd.DataFrame(market_rows)
                summary_df = pd.concat([summary_df, market_df], ignore_index=True)
            
            # Export to Excel - use openpyxl directly to avoid pandas issues
            self.add_log("📤 Exporting to Excel...", "INFO")
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill
                
                # Extract trading info
                symbol = df['Symbol'].iloc[0] if 'Symbol' in df.columns and len(df) > 0 else 'Unknown'
                period = df['Period'].iloc[0] if 'Period' in df.columns and len(df) > 0 else 'Unknown'
                company = df['Company'].iloc[0] if 'Company' in df.columns and len(df) > 0 else 'Unknown'
                currency = getattr(self, 'currency_var', None)
                currency = currency.get() if currency else 'USD'
                leverage = getattr(self, 'leverage_var', None)
                leverage = leverage.get() if leverage else '1:1'
                
                # Create workbook
                wb = Workbook()
                wb.remove(wb.active)
                
                # Write BACKTEST_DETAILS sheet
                backtest_details = pd.DataFrame({
                    "Category": ["Symbol", "Period", "Company", "Currency", "Leverage", "",
                                "Initial Deposit", "Final Balance", "Net Profit", "Return %", "",
                                "Total Trades", "Long Trades", "Short Trades", "Win Trades", "Loss Trades", "",
                                "Win Rate %", "Avg Profit per Trade", "Max Profit", "Max Loss",
                                "Gross Profit", "Gross Loss", "Profit Factor", "",
                                "Max Drawdown Absolute", "Max Drawdown %", "Recovery Factor", "AHPR %", "",
                                "Best Consecutive Wins", "Best Consecutive Losses"],
                    "Value": [symbol, period, company,
                             currency, leverage, "",
                             round(initial_balance, 2), round(closing_balance, 2), round(net_profit, 2),
                             round(return_percent, 2), "",
                             total_trades, len(buy_df), len(sell_df), win_count, loss_count, "",
                             round(winrate, 2), round(avg_profit, 2), round(max_profit, 2), round(max_loss, 2),
                             round(gross_profit, 2), round(gross_loss, 2), round(profit_factor, 2), "",
                             round(max_drawdown_absolute, 2), round(max_drawdown_relative, 2),
                             round(recovery_factor, 2), round(ahpr, 2), "", "N/A", "N/A"]
                })
                
                # Helper function to write DataFrame
                def write_dataframe_to_sheet(wb, df, sheet_name):
                    ws = wb.create_sheet(sheet_name)
                    for col_idx, col_name in enumerate(df.columns, 1):
                        ws.cell(row=1, column=col_idx, value=col_name)
                    for row_idx, row_data in enumerate(df.values, 2):
                        for col_idx, value in enumerate(row_data, 1):
                            ws.cell(row=row_idx, column=col_idx, value=value)
                    return ws
                
                # Write all sheets
                write_dataframe_to_sheet(wb, backtest_details, "BACKTEST_DETAILS")
                self.add_log(f"  ✓ BACKTEST_DETAILS", "INFO")
                
                write_dataframe_to_sheet(wb, df, "ALL_TRADES")
                self.add_log(f"  ✓ ALL_TRADES ({len(df)} trades)", "INFO")
                
                write_dataframe_to_sheet(wb, buy_df, "BUY_TRADES")
                self.add_log(f"  ✓ BUY_TRADES ({len(buy_df)} trades)", "INFO")
                
                write_dataframe_to_sheet(wb, sell_df, "SELL_TRADES")
                self.add_log(f"  ✓ SELL_TRADES ({len(sell_df)} trades)", "INFO")
                
                write_dataframe_to_sheet(wb, summary_df, "SUMMARY")
                self.add_log(f"  ✓ SUMMARY", "INFO")
                
                # Save workbook
                wb.save(output_file)
                self.add_log(f"  ✓ Workbook saved", "INFO")
                
            except Exception as e:
                self.add_log(f"✗ Export failed: {str(e)}", "ERROR")
                raise
            
            # Add charts sheet if images exist (after initial export)
            if chart_images:
                try:
                    wb = load_workbook(output_file)
                    
                    # Ensure all sheets are visible before adding charts
                    for sheet_name in wb.sheetnames:
                        wb[sheet_name].sheet_state = 'visible'
                    
                    for chart_image in chart_images:
                        if os.path.exists(chart_image):
                            # Create charts sheet if not exists
                            if "CHARTS" not in wb.sheetnames:
                                worksheet = wb.create_sheet("CHARTS", 0)
                            else:
                                worksheet = wb["CHARTS"]
                            
                            worksheet.sheet_state = 'visible'
                            img = XLImage(chart_image)
                            img.width = 800
                            img.height = 600
                            worksheet.add_image(img, 'A1')
                            self.add_log(f"  ✓ Added charts sheet", "INFO")
                            break
                    
                    # Ensure ALL sheets are visible after chart operations
                    for sheet_name in wb.sheetnames:
                        wb[sheet_name].sheet_state = 'visible'
                    
                    # Final safety: ensure at least one visible sheet
                    visible_sheets = [name for name in wb.sheetnames if wb[name].sheet_state == 'visible']
                    if not visible_sheets and wb.sheetnames:
                        wb[wb.sheetnames[0]].sheet_state = 'visible'
                    
                    wb.save(output_file)
                except PermissionError:
                    self.add_log(f"⚠️ Could not add charts - file may be open in Excel", "WARNING")
                except Exception as e:
                    self.add_log(f"⚠️ Could not add charts: {e}", "WARNING")
            
            # Format Excel
            self.add_log("🎨 Formatting Excel...", "INFO")
            
            wb = load_workbook(output_file)
            
            # Clean up temporary images
            for chart_image in chart_images:
                try:
                    if os.path.exists(chart_image):
                        os.remove(chart_image)
                except:
                    pass
            
            def format_profit_column(ws):
                for row in range(2, ws.max_row + 1):
                    cell = ws[f"{profit_col_letter}{row}"]
                    if cell.value is not None:
                        if cell.value > 0:
                            cell.font = Font(color="008000", bold=True)  # Green
                        elif cell.value < 0:
                            cell.font = Font(color="FF0000", bold=True)  # Red
            
            for sheet_name in ["ALL_TRADES", "BUY_TRADES", "SELL_TRADES"]:
                ws = wb[sheet_name]
                header = [cell.value for cell in ws[1]]
                if COL_PROFIT in header:
                    profit_col_letter = chr(header.index(COL_PROFIT) + 65)
                    format_profit_column(ws)
            
            # Format BACKTEST_DETAILS sheet
            if "BACKTEST_DETAILS" in wb.sheetnames:
                backtest_ws = wb["BACKTEST_DETAILS"]
                for cell in backtest_ws[1]:
                    cell.font = Font(bold=True)
                
                # Widen columns
                backtest_ws.column_dimensions['A'].width = 30
                backtest_ws.column_dimensions['B'].width = 20
            
            # Bold header summary
            summary_ws = wb["SUMMARY"]
            for cell in summary_ws[1]:
                cell.font = Font(bold=True)
            
            wb.save(output_file)
            self.add_log(f"  ✓ Excel formatting applied", "INFO")
            
            # Verify file was created
            if os.path.exists(output_file):
                file_size = os.path.getsize(output_file)
                self.add_log(f"✓ Excel file created: {file_size:,} bytes", "SUCCESS")
            else:
                raise FileNotFoundError(f"Excel file was not created at: {output_file}")
            
            self.add_log("=" * 60, "SUCCESS")
            self.add_log(f"✅ Conversion completed successfully!", "SUCCESS")
            self.add_log(f"📊 Output file: {output_file}", "SUCCESS")
            self.add_log("=" * 60, "SUCCESS")
            
            self.status_var.set(f"✓ Completed - Output: {output_file}")
            messagebox.showinfo("Success", f"Conversion completed!\n\nOutput: {output_file}")
            
        except Exception as e:
            error_msg = f"❌ Error: {str(e)}"
            self.add_log(error_msg, "ERROR")
            self.add_log("=" * 60, "ERROR")
            self.status_var.set("Error - Check logs")
            messagebox.showerror("Error", f"Conversion failed:\n\n{error_msg}")
        
        finally:
            self.is_converting = False
            self.root.after(0, lambda: self.convert_btn.config(state=tk.NORMAL))
    
    def convert_to_pdf(self):
        """Convert last generated Excel to PDF in background thread"""
        if not self.csv_file_var.get():
            messagebox.showerror("Error", "Please select a CSV file and convert to Excel first!")
            return
        
        output_folder = self.output_folder_var.get()
        csv_file = self.csv_file_var.get()
        input_basename = os.path.basename(os.path.splitext(csv_file)[0])
        excel_file = os.path.join(output_folder, input_basename + "_report.xlsx")
        
        if not os.path.exists(excel_file):
            messagebox.showerror("Error", 
                f"Excel file not found:\n\n{excel_file}\n\n" +
                f"Please:\n" +
                f"1. Make sure you clicked 'Convert to Excel' FIRST\n" +
                f"2. Check that output folder is correct\n" +
                f"3. Verify Excel file was created successfully")
            self.add_log(f"✗ PDF conversion failed: Excel file not found at {excel_file}", "ERROR")
            return
        
        if FPDF is None:
            messagebox.showerror("Error", "PDF library not installed!\n\nPlease install: pip install fpdf2")
            return
        
        # Disable PDF button and start conversion in thread
        self.pdf_btn.config(state=tk.DISABLED)
        thread = threading.Thread(target=self._do_pdf_conversion, args=(excel_file, output_folder), daemon=True)
        thread.start()
    
    def _generate_performance_charts(self, df, output_folder):
        """Generate performance trading charts as images (2 pages)"""
        if matplotlib is None:
            return []
        
        try:
            temp_images = []
            
            # Filter for trades with Saldo Akhir column
            if 'Saldo Awal' not in df.columns or 'Saldo Akhir' not in df.columns:
                return temp_images
            
            # Extract time info for analysis
            if 'Entry Time' in df.columns:
                df_copy = df.copy()
                df_copy['entry_datetime'] = pd.to_datetime(df_copy['Entry Time'], errors='coerce')
                df_copy['hour'] = df_copy['entry_datetime'].dt.hour
                df_copy['day_of_week'] = df_copy['entry_datetime'].dt.day_name()
                df_copy['month'] = df_copy['entry_datetime'].dt.month
            else:
                df_copy = df.copy()
            
            # PAGE 1: Equity Curve & Profit Distribution
            fig1, axes1 = plt.subplots(2, 1, figsize=(14, 10))
            fig1.suptitle('Trading Performance Analysis - Page 1', fontsize=16, fontweight='bold')
            
            # Equity Curve
            ax1 = axes1[0]
            balances = df['Saldo Akhir'].values
            ax1.plot(balances, linewidth=2.5, color='#00b0ff')
            ax1.fill_between(range(len(balances)), balances, alpha=0.3, color='#00b0ff')
            ax1.set_title('Equity Curve', fontweight='bold', fontsize=12)
            ax1.set_xlabel('Trade #', fontsize=10)
            ax1.set_ylabel('Balance ($)', fontsize=10)
            ax1.grid(True, alpha=0.3)
            
            # Profit Distribution
            ax2 = axes1[1]
            profits = df['Profit'].values
            colors = ['#2ecc71' if p > 0 else '#ff1744' for p in profits]
            ax2.bar(range(len(profits)), profits, color=colors, alpha=0.7, width=1)
            ax2.axhline(y=0, color='black', linestyle='-', linewidth=0.8)
            ax2.set_title('Profit per Trade', fontweight='bold', fontsize=12)
            ax2.set_xlabel('Trade #', fontsize=10)
            ax2.set_ylabel('Profit ($)', fontsize=10)
            ax2.grid(True, alpha=0.3, axis='y')
            
            plt.tight_layout()
            chart_image1 = os.path.join(output_folder, 'temp_chart_page1.png')
            fig1.savefig(chart_image1, dpi=150, bbox_inches='tight')
            plt.close(fig1)
            temp_images.append(chart_image1)
            
            # PAGE 2: Win/Loss & Profit by Market
            fig2, axes2 = plt.subplots(1, 2, figsize=(14, 6))
            fig2.suptitle('Trading Performance Analysis - Page 2', fontsize=16, fontweight='bold')
            
            # Win/Loss Pie Chart
            ax3 = axes2[0]
            win_count = len(df[df['Profit'] > 0])
            loss_count = len(df[df['Profit'] < 0])
            sizes = [win_count, loss_count]
            labels = [f'Win\n({win_count})', f'Loss\n({loss_count})']
            colors_pie = ['#2ecc71', '#ff1744']
            ax3.pie(sizes, labels=labels, colors=colors_pie, autopct='%1.1f%%', startangle=90, textprops={'fontsize': 11})
            ax3.set_title('Win/Loss Distribution', fontweight='bold', fontsize=12)
            
            # Profit by Market (if Market column exists)
            ax4 = axes2[1]
            if 'Market' in df.columns:
                market_profit = df.groupby('Market')['Profit'].sum().sort_values()
                colors_market = ['#2ecc71' if p > 0 else '#ff1744' for p in market_profit.values]
                market_profit.plot(kind='barh', ax=ax4, color=colors_market, alpha=0.7)
                ax4.set_title('Profit by Market', fontweight='bold', fontsize=12)
                ax4.set_xlabel('Total Profit ($)', fontsize=10)
                ax4.grid(True, alpha=0.3, axis='x')
            else:
                ax4.text(0.5, 0.5, 'No Market Data', ha='center', va='center', fontsize=12)
                ax4.set_title('Profit by Market', fontweight='bold', fontsize=12)
                ax4.axis('off')
            
            plt.tight_layout()
            chart_image2 = os.path.join(output_folder, 'temp_chart_page2.png')
            fig2.savefig(chart_image2, dpi=150, bbox_inches='tight')
            plt.close(fig2)
            temp_images.append(chart_image2)
            
            # PAGE 3: Profit/Loss by Hour & Weekday
            if 'entry_datetime' in df_copy.columns and df_copy['entry_datetime'].notna().any():
                fig3 = plt.figure(figsize=(14, 8))
                fig3.suptitle('Profit/Loss Analysis - Page 3', fontsize=16, fontweight='bold')
                gs = fig3.add_gridspec(1, 2, hspace=0.3, wspace=0.3)
                
                # Profit by hour
                ax_ph = fig3.add_subplot(gs[0, 0])
                hour_profit = df_copy.groupby('hour')['Profit'].sum()
                colors_hp = ['#2ecc71' if p > 0 else '#ff1744' for p in hour_profit.values]
                ax_ph.bar(hour_profit.index, hour_profit.values, color=colors_hp, alpha=0.7)
                ax_ph.axhline(y=0, color='black', linestyle='-', linewidth=0.8)
                ax_ph.set_title('Profit/Loss by Hour', fontweight='bold', fontsize=12)
                ax_ph.set_xlabel('Hour', fontsize=10)
                ax_ph.set_ylabel('Total Profit ($)', fontsize=10)
                ax_ph.grid(True, alpha=0.3, axis='y')
                
                # Profit by weekday
                ax_pw = fig3.add_subplot(gs[0, 1])
                day_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
                weekday_profit = df_copy.groupby('day_of_week')['Profit'].sum().reindex(day_order, fill_value=0)
                colors_pwp = ['#2ecc71' if p > 0 else '#ff1744' for p in weekday_profit.values]
                ax_pw.bar(range(len(weekday_profit)), weekday_profit.values, color=colors_pwp, alpha=0.7)
                ax_pw.axhline(y=0, color='black', linestyle='-', linewidth=0.8)
                ax_pw.set_title('Profit/Loss by Weekday', fontweight='bold', fontsize=12)
                ax_pw.set_xticks(range(len(weekday_profit)))
                ax_pw.set_xticklabels([d[:3] for d in weekday_profit.index], fontsize=10)
                ax_pw.set_ylabel('Total Profit ($)', fontsize=10)
                ax_pw.grid(True, alpha=0.3, axis='y')
                
                plt.tight_layout()
                chart_image3 = os.path.join(output_folder, 'temp_chart_page3.png')
                fig3.savefig(chart_image3, dpi=150, bbox_inches='tight')
                plt.close(fig3)
                temp_images.append(chart_image3)
                
                # PAGE 4: Entries & Profit by Month
                fig4 = plt.figure(figsize=(14, 8))
                fig4.suptitle('Monthly Analysis - Page 4', fontsize=16, fontweight='bold')
                gs4 = fig4.add_gridspec(1, 2, hspace=0.3, wspace=0.3)
                
                # Entries by month
                ax_em = fig4.add_subplot(gs4[0, 0])
                month_counts = df_copy['month'].value_counts().sort_index()
                ax_em.bar([f'Month {int(m)}' for m in month_counts.index], month_counts.values, color='#00b0ff', alpha=0.7)
                ax_em.set_title('Entries by Month', fontweight='bold', fontsize=12)
                ax_em.set_ylabel('Count', fontsize=10)
                ax_em.grid(True, alpha=0.3, axis='y')
                
                # Profit by month
                ax_pm = fig4.add_subplot(gs4[0, 1])
                month_profit = df_copy.groupby('month')['Profit'].sum().sort_index()
                colors_mp = ['#2ecc71' if p > 0 else '#ff1744' for p in month_profit.values]
                ax_pm.bar([f'Month {int(m)}' for m in month_profit.index], month_profit.values, color=colors_mp, alpha=0.7)
                ax_pm.axhline(y=0, color='black', linestyle='-', linewidth=0.8)
                ax_pm.set_title('Profit/Loss by Month', fontweight='bold', fontsize=12)
                ax_pm.set_ylabel('Total Profit ($)', fontsize=10)
                ax_pm.grid(True, alpha=0.3, axis='y')
                
                plt.tight_layout()
                chart_image4 = os.path.join(output_folder, 'temp_chart_page4.png')
                fig4.savefig(chart_image4, dpi=150, bbox_inches='tight')
                plt.close(fig4)
                temp_images.append(chart_image4)
            
            return temp_images
        
        except Exception as e:
            self.add_log(f"⚠️ Error generating charts: {e}", "WARNING")
            return []
    
    def _do_pdf_conversion(self, excel_file, output_folder):
        """Convert Excel to PDF (runs in background thread)"""
        try:
            self.add_log("=" * 60, "INFO")
            self.add_log("📄 Starting PDF conversion...", "INFO")
            self.add_log(f"Input: {excel_file}", "INFO")
            
            # Read Excel file
            self.add_log("📖 Reading Excel file...", "INFO")
            xls = pd.ExcelFile(excel_file)
            
            # Generate performance charts
            self.add_log("📊 Generating performance charts...", "INFO")
            all_trades_df = pd.read_excel(excel_file, sheet_name='ALL_TRADES')
            chart_images = self._generate_performance_charts(all_trades_df, output_folder)
            
            # Generate PDF filename (remove _report suffix if present)
            pdf_basename = os.path.basename(os.path.splitext(excel_file)[0])
            if pdf_basename.endswith('_report'):
                pdf_basename = pdf_basename[:-7]  # Remove '_report'
            pdf_file = os.path.join(output_folder, pdf_basename + ".pdf")
            
            # Create PDF with larger page size for better column fitting
            pdf = FPDF(orientation='L', unit='mm', format='A4')  # Landscape A4
            
            # Add charts page if available
            if chart_images:
                for chart_image in chart_images:
                    pdf.add_page()
                    pdf.set_font("Helvetica", 'B', 14)
                    pdf.cell(0, 10, "Performance Charts", ln=True)
                    pdf.ln(5)
                    
                    # Add chart image
                    if os.path.exists(chart_image):
                        try:
                            pdf.image(chart_image, x=10, y=30, w=280)
                            self.add_log(f"  ✓ Added performance charts", "INFO")
                        except Exception as e:
                            self.add_log(f"  ⚠️ Could not add chart image: {e}", "WARNING")
            
            # Add sheets (skip BACKTEST_DETAILS in PDF)
            for sheet_name in xls.sheet_names:
                # Skip BACKTEST_DETAILS from PDF
                if sheet_name == "BACKTEST_DETAILS":
                    continue
                
                self.add_log(f"  📄 Processing sheet: {sheet_name}", "INFO")
                
                # Read sheet data
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
                
                # Replace NaN values with empty string for display
                df = df.fillna('')
                
                if len(df) == 0:
                    continue
                
                # Add new page
                pdf.add_page()
                
                # Add title
                pdf.set_font("Helvetica", 'B', 14)
                pdf.cell(0, 10, f"Sheet: {sheet_name}", ln=True)
                pdf.ln(3)
                
                # Calculate optimal column widths
                num_cols = len(df.columns)
                available_width = pdf.w - 20  # Leave 10mm margin on each side
                
                # Estimate column widths (simple approach)
                col_widths = []
                min_col_width = 15  # Minimum column width in mm
                
                # Columns that need more width
                wide_columns = ['Entry Time', 'Exit Time', 'Reason', 'Entry Date', 'Exit Date', 'Metric']
                
                for col in df.columns:
                    # Get max string length in this column
                    max_len = len(str(col))  # Start with header
                    for val in df[col]:
                        max_len = max(max_len, len(str(val)))
                    
                    # Estimate width needed (roughly 0.5mm per character)
                    estimated_width = max(min_col_width, max_len * 0.4 + 2)
                    
                    # Add extra width for specific columns
                    if col == 'Metric':
                        estimated_width *= 2.0  # 100% wider (2x) for Metric column
                    elif col in wide_columns:
                        estimated_width *= 1.5  # 50% wider
                    
                    col_widths.append(estimated_width)
                
                # Scale columns to fit available width
                total_width = sum(col_widths)
                if total_width > available_width:
                    scale_factor = available_width / total_width
                    col_widths = [w * scale_factor for w in col_widths]
                
                # Set font for table
                pdf.set_font("Helvetica", 'B', 7)
                
                # Add table header with auto-fitted columns
                for col_idx, col in enumerate(df.columns):
                    col_width = col_widths[col_idx]
                    col_text = str(col)[:20]  # Truncate if too long
                    pdf.cell(col_width, 7, col_text, border=1, align='C')
                pdf.ln()
                
                # Add table data
                pdf.set_font("Helvetica", '', 6)
                for idx, row in df.iterrows():
                    for col_idx, col in enumerate(df.columns):
                        col_width = col_widths[col_idx]
                        value = str(row[col])
                        
                        # Truncate if too long
                        if len(value) > 25:
                            value = value[:22] + "..."
                        
                        # Align numeric values to right
                        align = 'R' if isinstance(row[col], (int, float)) else 'L'
                        pdf.cell(col_width, 5, value, border=1, align=align)
                    pdf.ln()
                
                self.add_log(f"  ✓ Added {sheet_name} ({len(df)} rows, {num_cols} columns)", "INFO")
            
            # Save PDF
            pdf.output(pdf_file)
            self.add_log(f"✓ PDF saved successfully", "SUCCESS")
            
            # Clean up temporary images
            for chart_image in chart_images:
                try:
                    if os.path.exists(chart_image):
                        os.remove(chart_image)
                except:
                    pass
            
            self.add_log("=" * 60, "SUCCESS")
            self.add_log(f"✅ PDF conversion completed!", "SUCCESS")
            self.add_log(f"📄 Output file: {pdf_file}", "SUCCESS")
            self.add_log("=" * 60, "SUCCESS")
            
            self.status_var.set(f"✓ PDF Ready - {pdf_file}")
            messagebox.showinfo("Success", f"PDF conversion completed!\n\nOutput: {pdf_file}")
            
        except Exception as e:
            error_msg = f"❌ Error: {str(e)}"
            self.add_log(error_msg, "ERROR")
            self.add_log("=" * 60, "ERROR")
            self.status_var.set("Error - Check logs")
            messagebox.showerror("Error", f"PDF conversion failed:\n\n{error_msg}")
        
        finally:
            self.root.after(0, lambda: self.pdf_btn.config(state=tk.NORMAL))


def main():
    root = tk.Tk()
    app = CSVToExcelConverter(root)
    root.mainloop()


if __name__ == "__main__":
    main()
